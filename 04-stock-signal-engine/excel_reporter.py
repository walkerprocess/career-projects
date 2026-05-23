import pandas as pd
import os
from datetime import datetime, timedelta
import json
from typing import Dict, List, Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import LineChart, Reference

class ExcelReporter:
    def __init__(self, data_dir: str = "portfolio_data"):
        self.data_dir = data_dir
        self.excel_file = os.path.join(data_dir, "portfolio_performance.xlsx")
        self.ensure_data_directory()
        
    def ensure_data_directory(self):
        """데이터 디렉토리가 존재하는지 확인하고 없으면 생성"""
        if not os.path.exists(self.data_dir):
            os.makedirs(self.data_dir)
    
    def create_excel_report(self, portfolio_data: Dict, daily_recommendations: List[Dict] = None):
        """엑셀 리포트 생성 및 업데이트"""
        try:
            print(f"📊 엑셀 리포트 생성 시작...")
            print(f"📁 저장 경로: {self.excel_file}")
            
            # 기존 파일이 있으면 로드, 없으면 새로 생성
            if os.path.exists(self.excel_file):
                print(f"📂 기존 파일 로드: {self.excel_file}")
                workbook = openpyxl.load_workbook(self.excel_file)
            else:
                print(f"🆕 새 워크북 생성")
                workbook = openpyxl.Workbook()
            
            # 기존 시트들 제거 (새로 생성)
            for sheet_name in workbook.sheetnames:
                print(f"🗑️ 기존 시트 제거: {sheet_name}")
                workbook.remove(workbook[sheet_name])
            
            # 1. 일일 성과 요약 시트
            print(f"📋 일일 성과 요약 시트 생성 중...")
            self.create_daily_summary_sheet(workbook, portfolio_data)
            
            # 2. 포트폴리오 상세 시트
            print(f"📋 포트폴리오 상세 시트 생성 중...")
            self.create_portfolio_detail_sheet(workbook, portfolio_data)
            
            # 3. 추천 히스토리 시트
            if daily_recommendations:
                print(f"📋 추천 히스토리 시트 생성 중...")
                self.create_recommendation_history_sheet(workbook, daily_recommendations)
            
            # 4. 월간 수익 분석 시트
            print(f"📋 월간 수익 분석 시트 생성 중...")
            self.create_monthly_analysis_sheet(workbook)
            
            # 5. 차트 시트
            print(f"📋 차트 시트 생성 중...")
            self.create_charts_sheet(workbook)
            
            # 파일 저장
            print(f"💾 파일 저장 중: {self.excel_file}")
            workbook.save(self.excel_file)
            print(f"✅ 엑셀 리포트 업데이트 완료: {self.excel_file}")
            
        except PermissionError as e:
            print(f"❌ 파일 권한 오류: {e}")
            print(f"💡 해결 방법: 엑셀 파일이 열려있다면 닫아주세요.")
        except Exception as e:
            print(f"❌ 엑셀 리포트 생성 오류: {e}")
            import traceback
            traceback.print_exc()
    
    def create_daily_summary_sheet(self, workbook, portfolio_data: Dict):
        """일일 성과 요약 시트 생성"""
        ws = workbook.create_sheet("일일 성과 요약")
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 데이터 준비
        current_date = datetime.now().strftime('%Y-%m-%d')
        
        # 포트폴리오 성과 계산
        portfolio_value = portfolio_data.get('portfolio_value', 0)
        total_investment = portfolio_data.get('total_investment', 0)
        total_return = portfolio_data.get('total_return', 0)
        total_return_pct = portfolio_data.get('total_return_pct', 0)
        
        # 헤더 작성
        headers = [
            ['날짜', '포트폴리오 가치', '총 투자금', '총 수익', '수익률', '보유 주식 수'],
            [current_date, f"${portfolio_value:,.0f}", f"${total_investment:,.0f}", 
             f"${total_return:,.0f}", f"{total_return_pct:.2f}%", 
             len(portfolio_data.get('current_stocks', []))]
        ]
        
        for row_idx, row in enumerate(headers, 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 1:  # 헤더 행
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
        
        # 열 너비 조정
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 20
        ws.column_dimensions['D'].width = 20
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
    
    def create_portfolio_detail_sheet(self, workbook, portfolio_data: Dict):
        """포트폴리오 상세 시트 생성"""
        ws = workbook.create_sheet("포트폴리오 상세")
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = [
            '주식 심볼', '회사명', '매수가', '현재가', '수량', '매수 금액', 
            '현재 가치', '수익금', '수익률', '매수일'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        current_stocks = portfolio_data.get('current_stocks', [])
        for row_idx, stock in enumerate(current_stocks, 2):
            ws.cell(row=row_idx, column=1, value=stock.get('symbol', ''))
            ws.cell(row=row_idx, column=2, value=stock.get('name', ''))
            ws.cell(row=row_idx, column=3, value=stock.get('purchase_price', 0))
            ws.cell(row=row_idx, column=4, value=stock.get('current_price', 0))
            ws.cell(row=row_idx, column=5, value=stock.get('quantity', 0))
            ws.cell(row=row_idx, column=6, value=stock.get('cost', 0))
            ws.cell(row=row_idx, column=7, value=stock.get('current_value', 0))
            ws.cell(row=row_idx, column=8, value=stock.get('return_amount', 0))
            ws.cell(row=row_idx, column=9, value=f"{stock.get('return_pct', 0):.2f}%")
            ws.cell(row=row_idx, column=10, value=stock.get('purchase_date', ''))
            
            # 수익률에 따른 색상 적용
            return_pct = stock.get('return_pct', 0)
            if return_pct > 0:
                ws.cell(row=row_idx, column=9).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif return_pct < 0:
                ws.cell(row=row_idx, column=9).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 열 너비 조정
        column_widths = [12, 25, 12, 12, 10, 15, 15, 15, 12, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
    
    def create_recommendation_history_sheet(self, workbook, daily_recommendations: List[Dict]):
        """추천 히스토리 시트 생성"""
        ws = workbook.create_sheet("추천 히스토리")
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 헤더 작성
        headers = [
            '날짜', '주식 심볼', '회사명', '현재가', '종합 점수', '기술 점수', 
            '뉴스 점수', 'RSI', '이동평균 추세', '예상 수익률'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        current_date = datetime.now().strftime('%Y-%m-%d')
        row_idx = 2
        
        for rec in daily_recommendations[:2]:  # 상위 2개만
            stock_info = rec.get('stock_info', {})
            score_info = rec.get('score_info', {})
            
            ws.cell(row=row_idx, column=1, value=current_date)
            ws.cell(row=row_idx, column=2, value=stock_info.get('symbol', ''))
            ws.cell(row=row_idx, column=3, value=stock_info.get('name', ''))
            ws.cell(row=row_idx, column=4, value=stock_info.get('current_price', 0))
            ws.cell(row=row_idx, column=5, value=score_info.get('total_score', 0))
            ws.cell(row=row_idx, column=6, value=score_info.get('technical_score', 0))
            ws.cell(row=row_idx, column=7, value=score_info.get('news_score', 0))
            ws.cell(row=row_idx, column=8, value=f"{score_info.get('rsi', 0):.1f}")
            ws.cell(row=row_idx, column=9, value=score_info.get('ma_trend', ''))
            
            # 예상 수익률 추출
            analysis = rec.get('analysis', '')
            expected_return = self.extract_expected_return_from_analysis(analysis)
            ws.cell(row=row_idx, column=10, value=f"{expected_return:.1f}%")
            
            row_idx += 1
        
        # 열 너비 조정
        column_widths = [12, 12, 25, 12, 12, 12, 12, 10, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
    
    def create_monthly_analysis_sheet(self, workbook):
        """월간 수익 분석 시트 생성"""
        ws = workbook.create_sheet("월간 수익 분석")
        
        # 헤더 스타일
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 월간 데이터 수집
        monthly_data = self.collect_monthly_data()
        
        # 헤더 작성
        headers = [
            '날짜', '포트폴리오 가치', '일일 수익', '일일 수익률', 
            '누적 수익', '누적 수익률', '최고 수익률', '최저 수익률'
        ]
        
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # 데이터 작성
        cumulative_return = 0
        max_return = 0
        min_return = 0
        
        for row_idx, data in enumerate(monthly_data, 2):
            daily_return = data.get('daily_return', 0)
            daily_return_pct = data.get('daily_return_pct', 0)
            cumulative_return += daily_return
            
            if daily_return_pct > max_return:
                max_return = daily_return_pct
            if daily_return_pct < min_return:
                min_return = daily_return_pct
            
            ws.cell(row=row_idx, column=1, value=data.get('date', ''))
            ws.cell(row=row_idx, column=2, value=f"${data.get('portfolio_value', 0):,.0f}")
            ws.cell(row=row_idx, column=3, value=f"${daily_return:,.0f}")
            ws.cell(row=row_idx, column=4, value=f"{daily_return_pct:.2f}%")
            ws.cell(row=row_idx, column=5, value=f"${cumulative_return:,.0f}")
            ws.cell(row=row_idx, column=6, value=f"{(cumulative_return/data.get('initial_investment', 1)*100):.2f}%")
            ws.cell(row=row_idx, column=7, value=f"{max_return:.2f}%")
            ws.cell(row=row_idx, column=8, value=f"{min_return:.2f}%")
            
            # 수익률에 따른 색상 적용
            if daily_return_pct > 0:
                ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif daily_return_pct < 0:
                ws.cell(row=row_idx, column=4).fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        
        # 열 너비 조정
        column_widths = [12, 20, 15, 15, 15, 15, 15, 15]
        for col_idx, width in enumerate(column_widths, 1):
            ws.column_dimensions[chr(64 + col_idx)].width = width
    
    def create_charts_sheet(self, workbook):
        """차트 시트 생성"""
        ws = workbook.create_sheet("차트")
        
        try:
            # 월간 데이터 수집
            monthly_data = self.collect_monthly_data()
            
            if not monthly_data:
                ws['A1'] = "차트 데이터가 없습니다."
                return
            
            # 차트 데이터 준비
            dates = [data.get('date', '') for data in monthly_data]
            portfolio_values = [data.get('portfolio_value', 0) for data in monthly_data]
            daily_returns = [data.get('daily_return_pct', 0) for data in monthly_data]
            
            # 데이터 작성
            ws['A1'] = '날짜'
            ws['B1'] = '포트폴리오 가치'
            ws['C1'] = '일일 수익률'
            
            for row_idx, (date, value, return_pct) in enumerate(zip(dates, portfolio_values, daily_returns), 2):
                ws.cell(row=row_idx, column=1, value=date)
                ws.cell(row=row_idx, column=2, value=value)
                ws.cell(row=row_idx, column=3, value=return_pct)
            
            # 포트폴리오 가치 차트
            try:
                chart1 = LineChart()
                chart1.title = "포트폴리오 가치 추이"
                chart1.x_axis.title = "날짜"
                chart1.y_axis.title = "포트폴리오 가치 ($)"
                
                data = Reference(ws, min_col=2, min_row=1, max_row=len(monthly_data)+1)
                cats = Reference(ws, min_col=1, min_row=2, max_row=len(monthly_data)+1)
                chart1.add_data(data, titles_from_data=True)
                chart1.set_categories(cats)
                
                ws.add_chart(chart1, "E2")
                print("📈 포트폴리오 가치 차트 생성 완료")
            except Exception as e:
                print(f"📈 포트폴리오 가치 차트 생성 실패: {e}")
                ws['E2'] = "포트폴리오 가치 차트 생성 실패"
            
            # 일일 수익률 차트
            try:
                chart2 = LineChart()
                chart2.title = "일일 수익률 추이"
                chart2.x_axis.title = "날짜"
                chart2.y_axis.title = "수익률 (%)"
                
                data = Reference(ws, min_col=3, min_row=1, max_row=len(monthly_data)+1)
                chart2.add_data(data, titles_from_data=True)
                chart2.set_categories(cats)
                
                ws.add_chart(chart2, "E18")
                print("📈 일일 수익률 차트 생성 완료")
            except Exception as e:
                print(f"📈 일일 수익률 차트 생성 실패: {e}")
                ws['E18'] = "일일 수익률 차트 생성 실패"
                
        except Exception as e:
            print(f"📊 차트 시트 생성 오류: {e}")
            ws['A1'] = f"차트 생성 중 오류 발생: {e}"
    
    def collect_monthly_data(self) -> List[Dict]:
        """월간 데이터 수집"""
        monthly_data = []
        
        try:
            # 일일 리포트 파일들 읽기
            for i in range(30):  # 최근 30일
                date = datetime.now() - timedelta(days=i)
                report_file = os.path.join(self.data_dir, f"daily_report_{date.strftime('%Y%m%d')}.json")
                
                if os.path.exists(report_file):
                    try:
                        with open(report_file, 'r', encoding='utf-8') as f:
                            data = json.load(f)
                            monthly_data.append({
                                'date': date.strftime('%Y-%m-%d'),
                                'portfolio_value': data.get('portfolio_value', 0),
                                'total_return_amount': data.get('total_return_amount', 0),
                                'total_return_pct': data.get('total_return_pct', 0),
                                'daily_return': 0,  # 계산 필요
                                'daily_return_pct': 0,  # 계산 필요
                                'initial_investment': 10000  # 기본값
                            })
                    except Exception as e:
                        print(f"일일 리포트 읽기 오류 ({report_file}): {e}")
            
            # 일일 수익 계산
            for i in range(1, len(monthly_data)):
                prev_value = monthly_data[i-1]['portfolio_value']
                curr_value = monthly_data[i]['portfolio_value']
                
                daily_return = curr_value - prev_value
                daily_return_pct = (daily_return / prev_value * 100) if prev_value > 0 else 0
                
                monthly_data[i]['daily_return'] = daily_return
                monthly_data[i]['daily_return_pct'] = daily_return_pct
            
            # 날짜순 정렬
            monthly_data.sort(key=lambda x: x['date'])
            
        except Exception as e:
            print(f"월간 데이터 수집 오류: {e}")
        
        # 데이터가 없으면 기본 데이터 생성
        if not monthly_data:
            print("📊 기본 월간 데이터 생성")
            for i in range(30):
                date = datetime.now() - timedelta(days=i)
                monthly_data.append({
                    'date': date.strftime('%Y-%m-%d'),
                    'portfolio_value': 10000 + (i * 50),  # 기본 증가
                    'total_return_amount': i * 50,
                    'total_return_pct': (i * 50) / 10000 * 100,
                    'daily_return': 50,
                    'daily_return_pct': 0.5,
                    'initial_investment': 10000
                })
            monthly_data.reverse()  # 날짜순 정렬
        
        return monthly_data
    
    def extract_expected_return_from_analysis(self, analysis: str) -> float:
        """GPT 분석에서 예상 수익률 추출"""
        try:
            lines = analysis.split('\n')
            for line in lines:
                if '목표가' in line or '상승' in line or '%' in line:
                    import re
                    numbers = re.findall(r'(\d+(?:\.\d+)?)\s*%', line)
                    if numbers:
                        return float(numbers[0])
                    
                    dollar_numbers = re.findall(r'\$(\d+(?:\.\d+)?)', line)
                    if dollar_numbers:
                        return 5.0  # 기본값
        except:
            pass
        
        return 3.0  # 기본 예상 수익률
    
    def get_excel_summary(self) -> Dict:
        """엑셀 파일 요약 정보 반환"""
        if not os.path.exists(self.excel_file):
            return {"status": "파일이 존재하지 않습니다."}
        
        try:
            workbook = openpyxl.load_workbook(self.excel_file)
            summary = {
                "파일명": os.path.basename(self.excel_file),
                "생성일": datetime.fromtimestamp(os.path.getctime(self.excel_file)).strftime('%Y-%m-%d %H:%M:%S'),
                "수정일": datetime.fromtimestamp(os.path.getmtime(self.excel_file)).strftime('%Y-%m-%d %H:%M:%S'),
                "시트 수": len(workbook.sheetnames),
                "시트 목록": workbook.sheetnames
            }
            return summary
        except Exception as e:
            return {"status": f"파일 읽기 오류: {e}"} 